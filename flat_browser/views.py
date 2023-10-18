from django.urls import reverse_lazy
from django.views.generic import ListView, FormView
from django.http import HttpResponse
from django.db.models import Q, Max, OuterRef, F
import pandas as pd
import io
from UliPlot.XLSX import auto_adjust_xlsx_column_width
from itertools import groupby
from datetime import datetime
from openpyxl.styles import Alignment
from django.core.cache import cache

from .models import Developer, Investment, Flat
from .forms import FlatForm


# Create your views here.

class DeveloperList(ListView):
    model = Developer
    context_object_name = "developer_list"


class InvestmentList(ListView):
    model = Investment
    context_object_name = "investment_list"
    template_name = "flat_browser/find_flat.html"

    def get_queryset(self):
        return Investment.objects.filter(developer__in=self.request.session['dev'])


class FlatList(ListView):
    model = Flat
    context_object_name = "flat_list"

    def get_queryset(self):
        investments = self.request.session['invest']
        flats = dict(self.request.GET)
        result_query = Flat.objects.filter(
            Q(floor__range=(int(*flats['floor_gte']), int(*flats['floor_lte']))) | Q(floor__isnull=True),
            Q(rooms__range=(int(*flats['rooms_gte']), int(*flats['rooms_lte']))) | Q(rooms__isnull=True),
            Q(price__range=(float(*flats['price_gte']), float(*flats['price_lte']))) | Q(price__isnull=True),
            Q(area__range=(float(*flats['area_gte']), float(*flats['area_lte']))) | Q(area__isnull=True),
            status__in=flats['status'],
            insertion_date=Flat.objects.latest("insertion_date").insertion_date,
            investment__in=investments).order_by("developer__name", "investment__name", "floor", "status", "area")
        return result_query

    def get(self, request, *args, **kwargs):
        self.request.session['flat_filter'] = dict(request.GET)
        return super().get(request, *args, **kwargs)


class FlatFormView(FormView):
    template_name = "flat_browser/find_flat.html"
    form_class = FlatForm
    success_url = reverse_lazy("flat_browser:flat")
    was_dev_chosen = None

    def get(self, request, *args, **kwargs):
        if request.GET.get('dev'):
            self.request.session['dev'] = dict(request.GET)['dev']
        if request.GET and not request.GET.get('dev'):
            form = FlatForm(request.GET)
            if form.is_valid():
                if request.GET.get('invest'):
                    request.session['invest'] = dict(request.GET)['invest']
                return self.form_valid(form)
            else:
                return self.form_invalid(form)
        else:

            return self.render_to_response(self.get_context_data())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['investment_list'] = InvestmentList.get_queryset(self)
        invest_ids = list(map(lambda x: x.id, context['investment_list']))
        self.request.session['invest'] = invest_ids
        return context


def export_excel_file(request):
    investments = request.session['invest']
    flats = request.session['flat_filter']
    query_flats = list(Flat.objects.filter(
        Q(floor__range=(int(*flats['floor_gte']), int(*flats['floor_lte']))) | Q(floor__isnull=True),
        Q(rooms__range=(int(*flats['rooms_gte']), int(*flats['rooms_lte']))) | Q(rooms__isnull=True),
        Q(price__range=(float(*flats['price_gte']), float(*flats['price_lte']))) | Q(price__isnull=True),
        Q(area__range=(float(*flats['area_gte']), float(*flats['area_lte']))) | Q(area__isnull=True),
        status__in=flats['status'],
        insertion_date=Flat.objects.latest("insertion_date").insertion_date,
        investment__in=investments).values("investment__name",
                                           "developer__name",
                                           "floor",
                                           "rooms",
                                           "area",
                                           "price",
                                           "status",
                                           "url",
                                           "insertion_date").order_by("developer__name"))
    sorted_flats = [{x: list(y)}
                    for x, y in groupby(query_flats, lambda z: z['developer__name'])]

    dev_names = list(map(lambda x: str(*x.keys()), sorted_flats))
    dataframes = [pd.DataFrame(items[1])
                  for flat in sorted_flats
                  for items in flat.items()
                  ]
    summaryDataframe = create_summary_to_excel(dataframes, dev_names)
    memory_file = io.BytesIO()
    with pd.ExcelWriter(memory_file, engine="openpyxl") as writer:
        summaryDataframe.to_excel(writer, sheet_name="Summary", startrow=1)
        auto_adjust_xlsx_column_width(summaryDataframe, writer, sheet_name="Summary", margin=1)
        writer.sheets['Summary'].merge_cells(start_row=1, start_column=1, end_row=1,
                                             end_column=10)
        writer.sheets['Summary'].cell(row=1, column=1).value = f"Creation datetime: " \
                                                               f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        writer.sheets['Summary'].cell(row=1, column=1).alignment = Alignment(horizontal='center')
        for name, df in zip(dev_names, dataframes):
            df.to_excel(writer, name)
            auto_adjust_xlsx_column_width(df, writer, sheet_name=name, margin=1)
    memory_file.seek(0)

    response = HttpResponse(memory_file,
                            headers={
                                "Content_Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                "Content-Disposition": 'attachment; filename="foo.xls"'
                            })
    return response


def create_summary_to_excel(dataframes: list, developer_names: list):
    summary = []
    for dataframe, developer_name in zip(dataframes, developer_names):
        summary.append({"developer": developer_name})
        investments = dataframe.groupby(by="investment__name")
        for investment in investments:
            currentInvestment = investments.get_group(investment[0])
            flatsStatus = currentInvestment['status'].value_counts(dropna=False).to_dict()
            if flatsStatus.get('wolne') is None:
                flatsStatus['wolne'] = 0
            if flatsStatus.get(None) is None:
                flatsStatus[None] = 0
            if flatsStatus.get('zarezerwowane') is None:
                flatsStatus['zarezerwowane'] = 0
            if flatsStatus.get('sprzedane') is None:
                flatsStatus['sprzedane'] = 0
            flatsToPricePerSqm = currentInvestment[
                currentInvestment['area'].notnull() & currentInvestment['status'].ne("sprzedane")]
            investmentSummary = {
                "investment_name": investment[0],
                "amount_of_all_flats": len(currentInvestment),
                "average_min_price_per_m2": round((flatsToPricePerSqm['price'] / flatsToPricePerSqm['area']).min(), 2),
                "average_max_price_per_m2": round((flatsToPricePerSqm['price'] / flatsToPricePerSqm['area']).max(), 2),
                "amount_of_free_flats": flatsStatus['wolne'] + flatsStatus[None],
                "amount_of_reserved_flats": flatsStatus.get('zarezerwowane'),
                "amount_of_sold_flats": flatsStatus.get('sprzedane'),
                "percentage_to_sold/all_flats": round(
                    (len(currentInvestment) - flatsStatus['wolne'] - flatsStatus[None] -
                     flatsStatus['zarezerwowane']) / len(currentInvestment), 2)
            }
            summary.append(investmentSummary)
    summaryDf = pd.DataFrame(summary)
    return summaryDf
